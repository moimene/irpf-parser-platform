"""
Excel M720 Exporter V2 — Genera XLSX con 5 hojas por clave BOE.

Convierte un M720DocumentExtraction en un archivo Excel con:
  - Hoja "Cuentas (C)": cuentas bancarias
  - Hoja "Valores (V)": acciones, ETFs, bonos
  - Hoja "IICs (I)": fondos de inversión
  - Hoja "Seguros (S)": pólizas de vida
  - Hoja "Inmuebles (B)": bienes inmuebles

Cada hoja tiene columnas mapeadas a los campos del Modelo 720 BOE.
Usa openpyxl (ya en dependencias) para máxima compatibilidad.

FORK LÓGICO: No comparte código con exporters V1.
"""

from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional

from app.schemas.m720_boe_v2 import (
    M720Cuenta,
    M720DocumentExtraction,
    M720IIC,
    M720Inmueble,
    M720Seguro,
    M720Valor,
)

logger = logging.getLogger(__name__)


def _addr_str(asset: object) -> str:
    """Serializa la dirección de la entidad a texto plano."""
    for attr in ("domicilio_entidad", "domicilio_inmueble"):
        addr = getattr(asset, attr, None)
        if addr is None:
            continue
        parts = [
            addr.calle or "",
            addr.poblacion or "",
            addr.provincia or "",
            addr.codigo_postal or "",
            addr.pais or "",
        ]
        return ", ".join(p for p in parts if p)
    return ""


def _base_row(asset: object) -> Dict[str, Any]:
    """Campos comunes de BaseM720Asset."""
    return {
        "País": getattr(asset, "pais_entidad_o_inmueble", ""),
        "Moneda": getattr(asset, "moneda_original", ""),
        "Condición": getattr(asset, "condicion_declarante", "Titular"),
        "Origen": getattr(asset, "origen_bien_derecho", "A"),
        "% Participación": getattr(asset, "porcentaje_participacion", 100.0),
        "NIF Representante": getattr(asset, "nif_representante", None) or "",
    }


def export_to_excel(extraction: M720DocumentExtraction) -> bytes:
    """
    Genera un archivo XLSX en memoria con 5 hojas.

    Returns:
        bytes del archivo XLSX listo para descargar.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Estilos
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def _create_sheet(
        ws: Any,
        title: str,
        columns: List[str],
        rows: List[Dict[str, Any]],
    ) -> None:
        """Configura una hoja con cabeceras y datos."""
        ws.title = title

        # Cabeceras
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Datos
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                # Formato numérico para importes
                if isinstance(value, float):
                    cell.number_format = "#,##0.00"

        # Auto-ajustar ancho de columnas
        for col_idx, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for row_idx in range(2, len(rows) + 2):
                val = str(ws.cell(row=row_idx, column=col_idx).value or "")
                max_len = max(max_len, min(len(val), 40))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

        # Filtro automático
        if rows:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    # ── Hoja 1: Cuentas (C) ──
    cuentas_cols = [
        "Subclave", "IBAN/Cuenta", "BIC", "Entidad", "NIF Entidad",
        "Dirección", "País", "Moneda", "Saldo 31/12", "Saldo Medio 4T",
        "F. Apertura", "F. Extinción", "Condición", "Origen", "% Participación",
    ]
    cuentas_rows = []
    for c in extraction.cuentas:
        row = _base_row(c)
        row.update({
            "Subclave": c.subclave,
            "IBAN/Cuenta": c.codigo_cuenta or "",
            "BIC": c.codigo_bic or "",
            "Entidad": c.denominacion_entidad or "",
            "NIF Entidad": c.nif_entidad or "",
            "Dirección": _addr_str(c),
            "Saldo 31/12": c.saldo_31_diciembre,
            "Saldo Medio 4T": c.saldo_medio_4T,
            "F. Apertura": c.fecha_apertura or "",
            "F. Extinción": c.fecha_extincion or "",
        })
        cuentas_rows.append(row)

    ws_cuentas = wb.active
    _create_sheet(ws_cuentas, "Cuentas (C)", cuentas_cols, cuentas_rows)

    # ── Hoja 2: Valores (V) ──
    valores_cols = [
        "Subclave", "ISIN", "Emisor", "NIF Entidad", "Dirección",
        "País", "Moneda", "Valor 31/12", "Nº Valores", "Representación",
        "F. Adquisición", "F. Transmisión", "Condición", "Origen", "% Participación",
    ]
    valores_rows = []
    for v in extraction.valores:
        row = _base_row(v)
        row.update({
            "Subclave": v.subclave,
            "ISIN": v.identificacion_valores or "",
            "Emisor": v.denominacion_entidad_emisora or "",
            "NIF Entidad": v.nif_entidad or "",
            "Dirección": _addr_str(v),
            "Valor 31/12": v.saldo_31_diciembre,
            "Nº Valores": v.numero_valores,
            "Representación": v.clave_representacion,
            "F. Adquisición": v.fecha_adquisicion or "",
            "F. Transmisión": v.fecha_transmision or "",
        })
        valores_rows.append(row)

    ws_valores = wb.create_sheet()
    _create_sheet(ws_valores, "Valores (V)", valores_cols, valores_rows)

    # ── Hoja 3: IICs (I) ──
    iics_cols = [
        "ISIN", "Gestora", "NIF Entidad", "Dirección",
        "País", "Moneda", "Valor Liquidativo 31/12", "Nº Participaciones",
        "F. Adquisición", "F. Transmisión", "Condición", "Origen", "% Participación",
    ]
    iics_rows = []
    for i in extraction.iics:
        row = _base_row(i)
        row.update({
            "ISIN": i.identificacion_valores or "",
            "Gestora": i.denominacion_entidad_gestora or "",
            "NIF Entidad": i.nif_entidad or "",
            "Dirección": _addr_str(i),
            "Valor Liquidativo 31/12": i.valor_liquidativo_31_diciembre,
            "Nº Participaciones": i.numero_valores,
            "F. Adquisición": i.fecha_adquisicion or "",
            "F. Transmisión": i.fecha_transmision or "",
        })
        iics_rows.append(row)

    ws_iics = wb.create_sheet()
    _create_sheet(ws_iics, "IICs (I)", iics_cols, iics_rows)

    # ── Hoja 4: Seguros (S) ──
    seguros_cols = [
        "Subclave", "Aseguradora", "NIF Entidad", "Dirección",
        "País", "Moneda", "Valor Rescate 31/12",
        "F. Contratación", "F. Extinción", "Condición", "Origen", "% Participación",
    ]
    seguros_rows = []
    for s in extraction.seguros:
        row = _base_row(s)
        row.update({
            "Subclave": s.subclave,
            "Aseguradora": s.denominacion_entidad_aseguradora or "",
            "NIF Entidad": s.nif_entidad or "",
            "Dirección": _addr_str(s),
            "Valor Rescate 31/12": s.valor_rescate_capitalizacion_31_diciembre,
            "F. Contratación": s.fecha_contratacion or "",
            "F. Extinción": s.fecha_extincion or "",
        })
        seguros_rows.append(row)

    ws_seguros = wb.create_sheet()
    _create_sheet(ws_seguros, "Seguros (S)", seguros_cols, seguros_rows)

    # ── Hoja 5: Inmuebles (B) ──
    inmuebles_cols = [
        "Clave Bien", "Subclave", "Tipo", "Registro", "Dirección Inmueble",
        "País", "Moneda", "Valor Adquisición", "Valor Transmisión",
        "F. Adquisición", "F. Transmisión", "Condición", "Origen", "% Participación",
    ]
    inmuebles_rows = []
    for b in extraction.inmuebles:
        row = _base_row(b)
        row.update({
            "Clave Bien": b.clave_bien,
            "Subclave": b.subclave,
            "Tipo": b.clave_tipo_inmueble,
            "Registro": b.denominacion_registro or "",
            "Dirección Inmueble": _addr_str(b),
            "Valor Adquisición": b.valor_adquisicion,
            "Valor Transmisión": b.valor_transmision,
            "F. Adquisición": b.fecha_adquisicion or "",
            "F. Transmisión": b.fecha_transmision or "",
        })
        inmuebles_rows.append(row)

    ws_inmuebles = wb.create_sheet()
    _create_sheet(ws_inmuebles, "Inmuebles (B)", inmuebles_cols, inmuebles_rows)

    # ── Guardar en memoria ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    total = (
        len(extraction.cuentas) + len(extraction.valores) + len(extraction.iics)
        + len(extraction.seguros) + len(extraction.inmuebles)
    )
    logger.info(
        "Excel V2 generado: %d activos en 5 hojas (C:%d V:%d I:%d S:%d B:%d), "
        "tamaño=%d bytes",
        total,
        len(extraction.cuentas),
        len(extraction.valores),
        len(extraction.iics),
        len(extraction.seguros),
        len(extraction.inmuebles),
        buffer.getbuffer().nbytes,
    )

    return buffer.getvalue()
