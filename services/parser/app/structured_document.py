import base64
import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from typing import List, Optional, Tuple

from app.extractors.base import extract_text_from_pdf
from app.schemas import (
    ParseDocumentRequest,
    SourceSpan,
    StructuredDocument,
    StructuredPage,
    StructuredRef,
    StructuredTable,
)


@dataclass
class StructuredLine:
    text: str
    page: int
    source_span: SourceSpan


def decode_content_bytes(content_base64: Optional[str]) -> Optional[bytes]:
    if not content_base64:
        return None
    try:
        return base64.b64decode(content_base64)
    except Exception:
        return None


def infer_source_type(request: ParseDocumentRequest) -> str:
    if request.source_type:
        return request.source_type

    normalized_mime = (request.mime_type or "").lower().strip()
    normalized_filename = request.filename.lower()

    if normalized_mime == "text/csv" or normalized_filename.endswith(".csv"):
        return "CSV"

    if normalized_mime in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    } or normalized_filename.endswith((".xlsx", ".xls")):
        return "XLSX"

    if normalized_mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or normalized_filename.endswith(".docx"):
        return "DOCX"

    if normalized_mime.startswith("image/") or normalized_filename.endswith(
        (".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".gif", ".webp")
    ):
        return "IMAGE"

    if request.text and request.text.strip():
        return "TEXT"

    return "PDF"


def build_structured_document(
    request: ParseDocumentRequest,
) -> Tuple[StructuredDocument, Optional[bytes], List[str]]:
    warnings: List[str] = []
    content_bytes = decode_content_bytes(request.content_base64)
    source_type = infer_source_type(request)

    if source_type == "CSV":
        structured_document, decode_warning = _build_csv_document(request, content_bytes)
        if decode_warning:
            warnings.append(decode_warning)
        return structured_document, content_bytes, warnings

    if source_type == "XLSX":
        structured_document, xlsx_warnings = _build_xlsx_document(content_bytes)
        warnings.extend(xlsx_warnings)
        return structured_document, content_bytes, warnings

    if source_type == "PDF":
        structured_document, pdf_warnings = _build_pdf_document(request, content_bytes)
        warnings.extend(pdf_warnings)
        return structured_document, content_bytes, warnings

    if source_type in {"IMAGE", "DOCX"}:
        if request.text and request.text.strip():
            return _build_text_document(request.text.strip(), source_type), content_bytes, warnings
        warnings.append(
            f"Formato {source_type} recibido sin capa estructurada activa; requiere OCR/Docling o revisión manual."
        )
        return (
            StructuredDocument(source_type=source_type, backend="unknown", pages=[], metadata={"page_count": 0}),
            content_bytes,
            warnings,
        )

    if request.text and request.text.strip():
        return _build_text_document(request.text.strip(), "TEXT"), content_bytes, warnings

    warnings.append("Documento sin contenido decodificable.")
    return (
        StructuredDocument(source_type="UNKNOWN", backend="unknown", pages=[], metadata={"page_count": 0}),
        content_bytes,
        warnings,
    )


def flatten_structured_text(document: StructuredDocument) -> str:
    chunks: List[str] = []
    for page in document.pages:
        page_text = (page.text or "").strip()
        if page_text:
            chunks.append(page_text)
            continue

        row_chunks: List[str] = []
        for table in page.tables:
            for row in [table.header, *table.rows]:
                row_text = _row_to_text(row)
                if row_text:
                    row_chunks.append(row_text)
        if row_chunks:
            chunks.append("\n".join(row_chunks))

    return "\n\n".join(chunk for chunk in chunks if chunk).strip()


def iter_structured_lines(document: StructuredDocument) -> List[StructuredLine]:
    lines: List[StructuredLine] = []
    seen: set[Tuple[int, str]] = set()

    for page in document.pages:
        for table in page.tables:
            header_text = _row_to_text(table.header)
            if header_text:
                key = (page.page, header_text)
                if key not in seen:
                    seen.add(key)
                    lines.append(
                        StructuredLine(
                            text=header_text,
                            page=page.page,
                            source_span=_make_table_span(
                                page=page.page,
                                table_id=table.table_id,
                                row=table.header,
                                kind="table_header",
                            ),
                        )
                    )

            for row_index, row in enumerate(table.rows):
                row_text = _row_to_text(row)
                if not row_text:
                    continue
                key = (page.page, row_text)
                if key not in seen:
                    seen.add(key)
                    lines.append(
                        StructuredLine(
                            text=row_text,
                            page=page.page,
                            source_span=_make_table_span(
                                page=page.page,
                                table_id=table.table_id,
                                row=row,
                                kind="table_row",
                                row_index=row_index,
                            ),
                        )
                    )

        for line in _iter_page_text_lines(page):
            key = (page.page, line.text)
            if key not in seen:
                seen.add(key)
                lines.append(line)

    return lines


def _build_text_document(text: str, source_type: str) -> StructuredDocument:
    return StructuredDocument(
        source_type=source_type,  # type: ignore[arg-type]
        backend="text",
        pages=[StructuredPage(page=1, text=text, tables=[])],
        metadata={"page_count": 1},
    )


def _build_csv_document(
    request: ParseDocumentRequest, content_bytes: Optional[bytes]
) -> Tuple[StructuredDocument, Optional[str]]:
    csv_text, decode_warning = _decode_text_payload(content_bytes, request.text)
    rows = _read_csv_rows(csv_text)
    header, body = _split_header_and_rows(rows)
    page_text = _rows_to_text(rows)
    table = StructuredTable(table_id="csv-1", page=1, source="csv:sheet-1", header=header, rows=body)
    return (
        StructuredDocument(
            source_type="CSV",
            backend="csv",
            pages=[StructuredPage(page=1, text=page_text, tables=[table] if rows else [])],
            metadata={
                "page_count": 1,
                "row_count": len(body),
                "column_count": max((len(row) for row in rows), default=0),
            },
        ),
        decode_warning,
    )


def _build_xlsx_document(content_bytes: Optional[bytes]) -> Tuple[StructuredDocument, List[str]]:
    warnings: List[str] = []
    if not content_bytes:
        warnings.append("Libro XLSX recibido sin contenido binario.")
        return (
            StructuredDocument(source_type="XLSX", backend="xlsx", pages=[], metadata={"page_count": 0}),
            warnings,
        )

    try:
        from openpyxl import load_workbook
    except ImportError:
        warnings.append("Dependencia openpyxl no disponible; no se puede estructurar XLSX.")
        return (
            StructuredDocument(source_type="XLSX", backend="unknown", pages=[], metadata={"page_count": 0}),
            warnings,
        )

    workbook = load_workbook(BytesIO(content_bytes), data_only=True, read_only=True)
    pages: List[StructuredPage] = []

    try:
        for index, worksheet in enumerate(workbook.worksheets, start=1):
            rows = [
                [_normalize_cell(cell) for cell in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            rows = [row for row in rows if any(cell is not None for cell in row)]
            header, body = _split_header_and_rows(rows)
            table = StructuredTable(
                table_id=f"xlsx-{index}",
                page=index,
                source=f"xlsx:{worksheet.title}",
                header=header,
                rows=body,
            )
            pages.append(
                StructuredPage(
                    page=index,
                    text=_rows_to_text(rows),
                    tables=[table] if rows else [],
                )
            )
    finally:
        workbook.close()

    return (
        StructuredDocument(
            source_type="XLSX",
            backend="xlsx",
            pages=pages,
            metadata={
                "page_count": len(pages),
                "sheet_count": len(workbook.sheetnames),
                "sheet_names": ", ".join(workbook.sheetnames),
            },
        ),
        warnings,
    )


def _build_pdf_document(
    request: ParseDocumentRequest, content_bytes: Optional[bytes]
) -> Tuple[StructuredDocument, List[str]]:
    warnings: List[str] = []
    if not content_bytes:
        if request.text and request.text.strip():
            return _build_text_document(request.text.strip(), "PDF"), warnings
        warnings.append("PDF recibido sin contenido binario.")
        return (
            StructuredDocument(source_type="PDF", backend="unknown", pages=[], metadata={"page_count": 0}),
            warnings,
        )

    pages_text, has_text = extract_text_from_pdf(content_bytes)
    pages: List[StructuredPage] = []
    backend = "text"

    try:
        import pdfplumber

        backend = "pdfplumber"
        with pdfplumber.open(BytesIO(content_bytes)) as pdf:
            if len(pages_text) < len(pdf.pages):
                pages_text = [(page.extract_text() or "") for page in pdf.pages]

            for index, page in enumerate(pdf.pages, start=1):
                raw_tables = page.extract_tables() or []
                structured_tables: List[StructuredTable] = []
                for table_index, raw_table in enumerate(raw_tables, start=1):
                    normalized_rows = [
                        [_normalize_cell(cell) for cell in row]
                        for row in raw_table
                        if row and any(cell is not None and str(cell).strip() for cell in row)
                    ]
                    if not normalized_rows:
                        continue

                    header, body = _split_header_and_rows(normalized_rows)
                    structured_tables.append(
                        StructuredTable(
                            table_id=f"pdf-{index}-{table_index}",
                            page=index,
                            source=f"pdf:page-{index}",
                            header=header,
                            rows=body,
                        )
                    )

                pages.append(
                    StructuredPage(
                        page=index,
                        text=pages_text[index - 1] if index - 1 < len(pages_text) else "",
                        tables=structured_tables,
                    )
                )
    except Exception as exc:
        warnings.append(f"No se pudieron estructurar tablas PDF con pdfplumber: {exc}")
        pages = [
            StructuredPage(page=index + 1, text=text, tables=[])
            for index, text in enumerate(pages_text)
        ]
        if not pages and request.text and request.text.strip():
            pages = [StructuredPage(page=1, text=request.text.strip(), tables=[])]

    return (
        StructuredDocument(
            source_type="PDF",
            backend=backend,  # type: ignore[arg-type]
            pages=pages,
            metadata={"page_count": len(pages), "native_text": has_text},
        ),
        warnings,
    )


def _decode_text_payload(content_bytes: Optional[bytes], fallback_text: Optional[str]) -> Tuple[str, Optional[str]]:
    if fallback_text and fallback_text.strip():
        return fallback_text.strip(), None

    if not content_bytes:
        return "", "Documento tabular sin contenido de texto."

    for encoding in ("utf-8-sig", "utf-8"):
        try:
            return content_bytes.decode(encoding), None
        except UnicodeDecodeError:
            continue

    try:
        return content_bytes.decode("latin-1"), "CSV decodificado con latin-1 por fallback."
    except UnicodeDecodeError:
        return "", "No se pudo decodificar el contenido tabular."


def _read_csv_rows(text: str) -> List[List[Optional[str]]]:
    if not text.strip():
        return []
    reader = csv.reader(StringIO(text))
    return [[_normalize_cell(cell) for cell in row] for row in reader]


def _normalize_cell(value: object) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()
    return text or None


def _split_header_and_rows(rows: List[List[Optional[str]]]) -> Tuple[List[Optional[str]], List[List[Optional[str]]]]:
    if not rows:
        return [], []
    if len(rows) == 1:
        return rows[0], []
    return rows[0], rows[1:]


def _rows_to_text(rows: List[List[Optional[str]]]) -> str:
    return "\n".join(
        row_text
        for row_text in (_row_to_text(row) for row in rows)
        if row_text
    )


def _row_to_text(row: List[Optional[str]]) -> str:
    return " ".join(cell for cell in row if cell).strip()


def _active_column_indices(row: List[Optional[str]]) -> List[int]:
    return [index for index, cell in enumerate(row) if cell is not None and str(cell).strip()]


def _make_table_span(
    *,
    page: int,
    table_id: str,
    row: List[Optional[str]],
    kind: str,
    row_index: Optional[int] = None,
) -> SourceSpan:
    row_text = _row_to_text(row)
    return SourceSpan(
        page=page,
        start=0,
        end=0,
        snippet=row_text or None,
        structured_ref=StructuredRef(
            kind=kind,  # type: ignore[arg-type]
            table_id=table_id,
            row_index=row_index,
            column_indices=_active_column_indices(row),
        ),
    )


def _iter_page_text_lines(page: StructuredPage) -> List[StructuredLine]:
    lines: List[StructuredLine] = []
    full_text = page.text or ""
    cursor = 0

    for line_index, raw_line in enumerate(full_text.splitlines()):
        stripped = raw_line.strip()
        if not stripped:
            cursor += len(raw_line) + 1
            continue

        leading_spaces = len(raw_line) - len(raw_line.lstrip())
        start = cursor + leading_spaces
        end = start + len(stripped)
        cursor += len(raw_line) + 1

        lines.append(
            StructuredLine(
                text=stripped,
                page=page.page,
                source_span=SourceSpan(
                    page=page.page,
                    start=max(0, start),
                    end=max(start, end),
                    snippet=stripped,
                    structured_ref=StructuredRef(
                        kind="page_text",
                        line_index=line_index,
                        column_indices=[],
                    ),
                ),
            )
        )

    return lines
