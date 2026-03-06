#!/usr/bin/env python3
"""
migrate_excel_to_supabase.py
============================
Migra los datos del Excel patrimonial (datosRPFFAGU2025.xlsx) a la tabla
irpf_patrimonio de Supabase.

Uso:
  SUPABASE_URL=https://xxx.supabase.co \
  SUPABASE_SERVICE_ROLE_KEY=eyJ... \
  CLIENT_ID=<uuid-del-cliente-en-irpf_clients> \
  python3 migrate_excel_to_supabase.py /ruta/al/datosRPFFAGU2025.xlsx

Requisitos:
  pip install openpyxl requests
"""

import os
import sys
import json
import re
import time
import requests
import openpyxl
from datetime import datetime, date, time as time_type

# ─── Configuración ────────────────────────────────────────────────────────────

SUPABASE_URL          = os.environ.get("SUPABASE_URL", "")
SUPABASE_SERVICE_KEY  = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
CLIENT_ID             = os.environ.get("CLIENT_ID", "")
BATCH_SIZE            = 50    # Reducido para evitar timeouts en hojas anchas
MAX_COLS              = 200   # Máximo de columnas por fila (truncar hojas muy anchas)
MAX_RETRIES           = 3     # Reintentos por batch

if not all([SUPABASE_URL, SUPABASE_SERVICE_KEY, CLIENT_ID]):
    print("ERROR: Faltan variables de entorno SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY o CLIENT_ID")
    sys.exit(1)

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else "datosRPFFAGU2025.xlsx"

# ─── Mapeo de hojas a categorías ──────────────────────────────────────────────

SHEET_CATEGORY_MAP = {
    # Inventario / Posiciones
    "INVENTARIO": "inventario",
    "ACCIONES": "inventario",
    "CARTERA": "inventario",
    # Goldman Sachs
    "GS": "goldman",
    "GOLDMAN": "goldman",
    "GOLDMAN SACHS": "goldman",
    # Citi
    "CITI": "citi",
    "CITIBANK": "citi",
    "CITIGROUP": "citi",
    "BROKERAGE": "citi",
    # J.P. Morgan
    "JPM": "jpmorgan",
    "JPMORGAN": "jpmorgan",
    "JP MORGAN": "jpmorgan",
    # Pictet
    "PICTET": "pictet",
    # Derivados
    "FORWARDS": "derivados",
    "DERIVADOS": "derivados",
    "OPCIONES": "derivados",
    # Inmuebles
    "INMUEBLES": "inmuebles",
    "ESPALTER": "inmuebles",
    "JOSELITO": "inmuebles",
    "TATO": "inmuebles",
    # Obras de Arte
    "ARTE": "obras_arte",
    "OBRAS": "obras_arte",
    "COLECCION": "obras_arte",
    # Private Equity
    "PE": "private_equity",
    "PRIVATE EQUITY": "private_equity",
    "FONDOS": "private_equity",
    # Tipos de cambio
    "TC": "tipos_cambio",
    "TIPOS": "tipos_cambio",
    "CAMBIO": "tipos_cambio",
    "FX": "tipos_cambio",
}

def detect_category(sheet_name: str) -> str:
    """Detecta la categoría patrimonial a partir del nombre de la hoja."""
    upper = sheet_name.upper()
    for key, cat in SHEET_CATEGORY_MAP.items():
        if key in upper:
            return cat
    return "inventario"  # fallback

def detect_ejercicio(sheet_name: str) -> int | None:
    """Extrae el año fiscal del nombre de la hoja si está presente."""
    match = re.search(r"(20\d{2})", sheet_name)
    return int(match.group(1)) if match else None

def serialize_value(v):
    """Convierte valores de celda a tipos JSON serializables."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time_type):
        return v.strftime("%H:%M:%S")
    if isinstance(v, float):
        if v != v:  # NaN
            return None
        if v == float('inf') or v == float('-inf'):
            return None
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, bool):
        return v
    return str(v) if v is not None else None

def read_sheet(ws, max_cols: int = MAX_COLS) -> tuple[list[str], list[dict]]:
    """Lee una hoja y devuelve (columnas, filas_como_dict).
    Trunca columnas si hay más de max_cols para evitar payloads gigantes.
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    # Primera fila no vacía como cabecera
    header_idx = 0
    for i, row in enumerate(rows):
        if any(c is not None for c in row):
            header_idx = i
            break

    raw_headers = rows[header_idx]

    # Truncar si hay demasiadas columnas
    if len(raw_headers) > max_cols:
        raw_headers = raw_headers[:max_cols]

    headers = []
    seen = {}
    for i, h in enumerate(raw_headers):
        if h is None or str(h).strip() == "":
            col_name = f"col_{i}"
        else:
            col_name = str(h).strip()[:100]  # Limitar longitud del nombre
        # Desduplicar nombres de columna
        if col_name in seen:
            seen[col_name] += 1
            col_name = f"{col_name}_{seen[col_name]}"
        else:
            seen[col_name] = 0
        headers.append(col_name)

    records = []
    for row in rows[header_idx + 1:]:
        if not any(c is not None for c in row):
            continue  # Saltar filas completamente vacías
        # Truncar fila si tiene más columnas que las cabeceras
        row_truncated = row[:len(headers)]
        record = {}
        for col, val in zip(headers, row_truncated):
            record[col] = serialize_value(val)
        records.append(record)

    return headers, records

def supabase_insert_batch(table: str, rows: list[dict], retry: int = 0) -> int:
    """Inserta un batch de filas en Supabase vía REST API con reintentos."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    try:
        resp = requests.post(url, headers=headers, json=rows, timeout=60)
        if resp.status_code in (200, 201):
            return len(rows)
        elif resp.status_code in (520, 521, 522, 524) and retry < MAX_RETRIES:
            # Cloudflare errors — esperar y reintentar
            wait = 2 ** retry
            print(f"    Cloudflare error {resp.status_code}, reintentando en {wait}s...")
            time.sleep(wait)
            return supabase_insert_batch(table, rows, retry + 1)
        else:
            raise RuntimeError(f"Supabase error {resp.status_code}: {resp.text[:200]}")
    except requests.exceptions.ReadTimeout:
        if retry < MAX_RETRIES:
            wait = 5 * (retry + 1)
            print(f"    Timeout, reintentando en {wait}s con batch más pequeño...")
            time.sleep(wait)
            # Dividir el batch a la mitad
            mid = len(rows) // 2
            if mid == 0:
                raise RuntimeError("Batch de 1 elemento sigue dando timeout")
            inserted = 0
            inserted += supabase_insert_batch(table, rows[:mid], retry + 1)
            inserted += supabase_insert_batch(table, rows[mid:], retry + 1)
            return inserted
        raise

def upsert_hoja(client_id: str, categoria_id: str, nombre: str,
                ejercicio: int | None, columnas: list[str],
                num_filas: int, kpis: dict) -> None:
    """Registra los metadatos de la hoja en irpf_hojas."""
    url = f"{SUPABASE_URL}/rest/v1/irpf_hojas"
    headers = {
        "apikey": SUPABASE_SERVICE_KEY,
        "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    payload = {
        "client_id":    client_id,
        "categoria_id": categoria_id,
        "nombre":       nombre,
        "ejercicio":    ejercicio,
        "num_filas":    num_filas,
        "columnas":     columnas[:100],  # Limitar columnas en metadatos
        "kpis":         kpis,
    }
    try:
        resp = requests.post(url, headers=headers, json=payload, timeout=30)
        if resp.status_code not in (200, 201):
            print(f"  WARN hoja upsert: {resp.status_code} {resp.text[:100]}")
    except Exception as e:
        print(f"  WARN hoja upsert error: {e}")

# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    print(f"Abriendo {EXCEL_PATH} ...")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    sheets = wb.sheetnames
    print(f"  {len(sheets)} hojas encontradas")

    # Verificar cuántos registros ya existen (para reanudar si es necesario)
    check_resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/irpf_patrimonio?client_id=eq.{CLIENT_ID}&select=hoja",
        headers={
            "apikey": SUPABASE_SERVICE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
            "Prefer": "count=exact",
            "Range": "0-0",
        }
    )
    existing_count = int(check_resp.headers.get("Content-Range", "0/0").split("/")[-1])
    print(f"  Registros existentes en Supabase: {existing_count}")

    # Obtener hojas ya migradas
    hojas_resp = requests.get(
        f"{SUPABASE_URL}/rest/v1/irpf_hojas?client_id=eq.{CLIENT_ID}&select=nombre",
        headers={
            "apikey": SUPABASE_SERVICE_KEY,
            "Authorization": f"Bearer {SUPABASE_SERVICE_KEY}",
        }
    )
    hojas_migradas = set()
    if hojas_resp.status_code == 200:
        hojas_migradas = {h["nombre"] for h in hojas_resp.json()}
    print(f"  Hojas ya migradas: {len(hojas_migradas)}")

    total_inserted = 0
    total_sheets   = 0
    errors         = []

    for sheet_name in sheets:
        # Saltar hojas ya migradas
        if sheet_name in hojas_migradas:
            print(f"\n[{sheet_name}] → ya migrada, omitiendo")
            continue

        ws = wb[sheet_name]
        categoria_id = detect_category(sheet_name)
        ejercicio    = detect_ejercicio(sheet_name)

        print(f"\n[{sheet_name}] → {categoria_id} (ejercicio: {ejercicio})")

        try:
            columnas, records = read_sheet(ws)
        except Exception as e:
            print(f"  ERROR leyendo hoja: {e}")
            errors.append(f"{sheet_name}: error lectura - {e}")
            continue

        if not records:
            print(f"  Sin datos — omitida")
            # Registrar hoja vacía para no reintentar
            upsert_hoja(CLIENT_ID, categoria_id, sheet_name, ejercicio, [], 0, {})
            continue

        num_cols = len(columnas)
        if num_cols > MAX_COLS:
            print(f"  AVISO: {num_cols} columnas → truncando a {MAX_COLS}")

        print(f"  {len(records)} filas, {min(num_cols, MAX_COLS)} columnas")

        # Preparar rows para irpf_patrimonio
        rows_to_insert = []
        for i, rec in enumerate(records):
            rows_to_insert.append({
                "client_id":    CLIENT_ID,
                "categoria_id": categoria_id,
                "hoja":         sheet_name,
                "ejercicio":    ejercicio,
                "fila":         i + 2,  # +2 por la cabecera y base 1
                "datos":        rec,
                "fuente":       "excel_manual",
            })

        # Insertar en batches
        inserted = 0
        for i in range(0, len(rows_to_insert), BATCH_SIZE):
            batch = rows_to_insert[i:i + BATCH_SIZE]
            try:
                n = supabase_insert_batch("irpf_patrimonio", batch)
                inserted += n
            except Exception as e:
                print(f"  ERROR batch {i//BATCH_SIZE + 1}: {e}")
                errors.append(f"{sheet_name} batch {i//BATCH_SIZE + 1}: {e}")

        # Calcular KPIs básicos (sumas de columnas numéricas)
        kpis = {}
        for col in columnas[:10]:  # Solo primeras 10 columnas para KPIs
            vals = [r.get(col) for r in records if isinstance(r.get(col), (int, float))]
            if vals:
                kpis[col] = {"sum": round(sum(vals), 4), "count": len(vals)}

        # Registrar metadatos de la hoja
        upsert_hoja(CLIENT_ID, categoria_id, sheet_name, ejercicio,
                    columnas, len(records), kpis)

        print(f"  Insertadas {inserted}/{len(records)} filas")
        total_inserted += inserted
        total_sheets   += 1

        # Pequeña pausa para no saturar la API
        time.sleep(0.1)

    wb.close()
    print(f"\n{'='*50}")
    print(f"Migracion completada: {total_sheets} hojas nuevas, {total_inserted} registros insertados")
    print(f"Cliente ID: {CLIENT_ID}")
    if errors:
        print(f"\nErrores ({len(errors)}):")
        for e in errors[:10]:
            print(f"  - {e}")

if __name__ == "__main__":
    main()
